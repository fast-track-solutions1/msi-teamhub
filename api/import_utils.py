# api/import_utils.py - LOGIQUE D'IMPORTATION GÉNÉRIQUE - ✅ COMPLET

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging
from django.apps import apps
from django.db import transaction
from datetime import datetime

logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION DES MODÈLES IMPORTABLES
# ============================================================================

IMPORTABLE_MODELS = {
    'societe': {
        'app': 'api',
        'model': 'Societe',
        'name': 'Société',
        'unique_field': 'nom',
        'exclude_fields': ['id', 'date_creation'],
    },
    'departement': {
        'app': 'api',
        'model': 'Departement',
        'name': 'Département',
        'unique_field': 'numero',
        'exclude_fields': ['id', 'date_creation'],
    },
    'circuit': {
        'app': 'api',
        'model': 'Circuit',
        'name': 'Circuit',
        'unique_field': None,
        'exclude_fields': ['id', 'date_creation'],
    },
    'service': {
        'app': 'api',
        'model': 'Service',
        'name': 'Service',
        'unique_field': 'nom',
        'exclude_fields': ['id', 'date_creation'],
    },
    'grade': {
        'app': 'api',
        'model': 'Grade',
        'name': 'Grade',
        'unique_field': 'nom',
        'exclude_fields': ['id', 'date_creation'],
    },
    'creneau_travail': {
        'app': 'api',
        'model': 'CreneauTravail',
        'name': 'Créneau de Travail',
        'unique_field': 'nom',
        'exclude_fields': ['id', 'date_creation'],
    },
    'type_acces': {
        'app': 'api',
        'model': 'TypeAcces',
        'name': 'Type d\'Accès',
        'unique_field': 'nom',
        'exclude_fields': ['id'],
    },
    'outil_travail': {
        'app': 'api',
        'model': 'OutilTravail',
        'name': 'Outil de Travail',
        'unique_field': 'nom',
        'exclude_fields': ['id'],
    },
    'equipement': {
        'app': 'api',
        'model': 'Equipement',
        'name': 'Équipement',
        'unique_field': None,
        'exclude_fields': ['id', 'date_creation'],
    },
    'salarie': {
        'app': 'api',
        'model': 'Salarie',
        'name': 'Salarié',
        'unique_field': 'matricule',
        'exclude_fields': ['id', 'date_creation', 'date_modification'],
    },
        'accesapplication': {'app': 'api', 'model': 'AccesApplication', 'name': 'Accès Application', 'unique_field': None, 'exclude_fields': ['id', 'date_creation']},
    'equipementinstance': {'app': 'api', 'model': 'EquipementInstance', 'name': 'Équipement Instance', 'unique_field': 'numero_serie', 'exclude_fields': ['id', 'date_creation']},
    'horairesalarie': {'app': 'api', 'model': 'HoraireSalarie', 'name': 'Horaire Salarié', 'unique_field': None, 'exclude_fields': ['id', 'date_creation']},
}

# ============================================================================
# FONCTION POUR LISTER LES MODÈLES IMPORTABLES
# ============================================================================

def get_importable_models():
    """Retourne le dictionnaire des modèles importables"""
    return IMPORTABLE_MODELS

# ============================================================================
# CLASSE GÉNÉRIQUE D'IMPORTATION
# ============================================================================

class GenericImporter:
    """Classe générique pour importer n'importe quel modèle Django depuis Excel"""

    def __init__(self, model_name: str):
        """
        Initialise l'importeur
        
        Args:
            model_name: clé du modèle à importer (ex: 'salarie', 'departement')
        """
        if model_name not in IMPORTABLE_MODELS:
            raise ValueError(f"Modèle '{model_name}' non importable. Disponibles: {list(IMPORTABLE_MODELS.keys())}")
        
        self.model_name = model_name
        self.config = IMPORTABLE_MODELS[model_name]
        self.Model = apps.get_model(self.config['app'], self.config['model'])
        self.results = {
            'inserted': 0,
            'updated': 0,
            'errors': [],
            'warnings': []
        }

    def get_model_structure(self) -> dict:
        """
        Retourne la structure du modèle (champs, types, obligatoires, etc.)
        
        Returns:
            dict: Structure du modèle avec info sur ForeignKey
        """
        try:
            fields = self._get_importable_fields()
            structure = {
                'fields': [],
                'unique_field': self.config.get('unique_field'),
                'exclude_fields': self.config.get('exclude_fields', [])
            }
            
            for field_name in fields:
                field = self.Model._meta.get_field(field_name)
                field_info = {
                    'name': field_name,
                    'type': field.get_internal_type(),
                    'required': not field.null and not field.blank,
                }
                
                # Si c'est une ForeignKey, ajouter info sur la relation
                if field.get_internal_type() == 'ForeignKey':
                    related_model = field.related_model
                    field_info['is_foreign_key'] = True
                    field_info['related_model'] = related_model.__name__
                    # Récupérer les valeurs possibles
                    try:
                        values = list(related_model.objects.values_list('nom', flat=True))
                        field_info['possible_values'] = values
                    except:
                        field_info['possible_values'] = []
                else:
                    field_info['is_foreign_key'] = False
                
                structure['fields'].append(field_info)
            
            return structure
        except Exception as e:
            logger.error(f"Erreur lors de la récupération de la structure: {str(e)}")
            raise

    def _get_importable_fields(self) -> list:
        """
        Retourne la liste des champs importables du modèle
        
        Returns:
            list: Noms des champs importables
        """
        exclude = self.config.get('exclude_fields', [])
        return [f.name for f in self.Model._meta.get_fields() if f.name not in exclude and not f.many_to_many]

    def generate_template(self) -> bytes:
        """
        Génère un fichier Excel template basé sur les champs du modèle
        Inclut: en-têtes formatés, feuille d'instructions, listes déroulantes pour FK
        
        Returns:
            bytes: Fichier Excel en bytes
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Données"
            
            # Récupérer tous les champs
            fields = self._get_importable_fields()
            
            # Style pour l'en-tête
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Ajouter les en-têtes
            for col_num, field_name in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = field_name
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Ajouter 10 lignes vides d'exemple
            for row in range(2, 12):
                for col in range(1, len(fields) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
            
            # Ajuster la largeur des colonnes
            for col_num, field_name in enumerate(fields, 1):
                ws.column_dimensions[get_column_letter(col_num)].width = max(15, len(field_name) + 5)
            
            # Figer l'en-tête
            ws.freeze_panes = "A2"
            
            # AJOUTER LES LISTES DÉROULANTES POUR LES FOREIGNKEY
            for col_num, field_name in enumerate(fields, 1):
                try:
                    field = self.Model._meta.get_field(field_name)
                    if field.get_internal_type() == 'ForeignKey':
                        related_model = field.related_model
                        # Récupérer les valeurs possibles
                        values = list(related_model.objects.values_list('nom', flat=True).order_by('nom'))
                        if values:
                            # Créer la liste déroulante
                            dv = DataValidation(type='list', formula1=f'"{",".join(map(str, values))}"', allow_blank=True)
                            dv.error = f'Sélectionnez une valeur valide pour {field_name}'
                            dv.errorTitle = f'Valeur invalide: {field_name}'
                            ws.add_data_validation(dv)
                            
                            # Appliquer la liste déroulante à la colonne (lignes 2 à 1000)
                            col_letter = get_column_letter(col_num)
                            for row in range(2, 1001):
                                dv.add(f'{col_letter}{row}')
                except Exception as e:
                    logger.warning(f"Impossible d'ajouter liste déroulante pour {field_name}: {str(e)}")
            
            # CRÉER FEUILLE D'INSTRUCTIONS
            instructions_sheet = wb.create_sheet("Instructions")
            instructions_sheet['A1'] = "INSTRUCTIONS D'IMPORT"
            instructions_sheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            instructions_sheet['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            row = 3
            instructions_sheet[f'A{row}'] = "Colonnes obligatoires:"
            instructions_sheet[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            unique_field = self.config.get('unique_field')
            for field_name in fields:
                try:
                    field = self.Model._meta.get_field(field_name)
                    is_required = not field.null and not field.blank
                    is_unique = field_name == unique_field
                    
                    if is_required or is_unique:
                        marker = "⚠️ " if is_required else "🔑 "
                        instructions_sheet[f'A{row}'] = f"{marker} {field_name}"
                        if is_unique:
                            instructions_sheet[f'B{row}'] = "(Clé unique)"
                        row += 1
                except:
                    pass
            
            row += 2
            instructions_sheet[f'A{row}'] = "Champs avec sélection (dropdown):"
            instructions_sheet[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            for field_name in fields:
                try:
                    field = self.Model._meta.get_field(field_name)
                    if field.get_internal_type() == 'ForeignKey':
                        related_model = field.related_model
                        values = list(related_model.objects.values_list('nom', flat=True).order_by('nom'))
                        instructions_sheet[f'A{row}'] = f"• {field_name}"
                        instructions_sheet[f'B{row}'] = ", ".join(map(str, values[:5]))  # Afficher les 5 premiers
                        if len(values) > 5:
                            instructions_sheet[f'B{row}'] = instructions_sheet[f'B{row}'].value + f"... (+{len(values)-5} autres)"
                        row += 1
                except:
                    pass
            
            # Ajuster la largeur des colonnes d'instructions
            instructions_sheet.column_dimensions['A'].width = 30
            instructions_sheet.column_dimensions['B'].width = 80
            
            # Sauvegarder dans BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error(f"Erreur lors de la génération du template: {str(e)}")
            raise

    def import_from_excel(self, file) -> dict:
        """
        Importe les données depuis un fichier Excel
        
        Args:
            file: Fichier Excel uploadé
            
        Returns:
            dict: Résultat de l'import {inserted, updated, errors, warnings}
        """
        try:
            # Lire le fichier Excel
            df = pd.read_excel(file)
            
            if df.empty:
                raise ValueError("Le fichier Excel est vide")
            
            # Nettoyer les données
            df = df.fillna('')  # Remplacer NaN par chaîne vide
            df.columns = [col.strip().lower().replace(' ', '_') for col in df.columns]  # Normaliser les colonnes
            
            logger.info(f"Import de {len(df)} lignes pour {self.model_name}")
            
            # Importer chaque ligne
            with transaction.atomic():
                for idx, row in df.iterrows():
                    try:
                        self._import_row(row, idx + 2)  # +2 car idx commence à 0 et ligne 1 est l'en-tête
                    except Exception as e:
                        self.results['errors'].append({
                            'row': idx + 2,
                            'error': str(e)
                        })
                        logger.error(f"Erreur ligne {idx + 2}: {str(e)}")
            
            return self.results
        except Exception as e:
            logger.error(f"Erreur lors de l'import: {str(e)}")
            self.results['errors'].append({'row': 0, 'error': f"Erreur générale: {str(e)}"})
            return self.results

    def _import_row(self, row: pd.Series, row_num: int):
        """
        Importe une ligne unique
        
        Args:
            row: Série pandas représentant la ligne
            row_num: Numéro de la ligne (pour les erreurs)
        """
        # Préparer les données
        data = {}
        for key, value in row.items():
            if pd.isna(value) or value == '':
                continue
            
            # Valider et convertir le champ
            data[key] = self._convert_field_value(key, value)
        
        if not data:
            self.results['warnings'].append({'row': row_num, 'warning': 'Ligne vide'})
            return
        
        # Déterminer clé unique
        unique_field = self.config.get('unique_field')
        update_or_create_kwargs = {}
        
        if unique_field and unique_field in data:
            update_or_create_kwargs[unique_field] = data[unique_field]
        elif 'id' in data and data['id']:
            update_or_create_kwargs['id'] = data['id']
        else:
            # Aucune clé unique, créer directement
            obj = self.Model.objects.create(**data)
            self.results['inserted'] += 1
            return
        
        # Update or create
        defaults = {k: v for k, v in data.items() if k not in update_or_create_kwargs}
        obj, created = self.Model.objects.update_or_create(
            **update_or_create_kwargs,
            defaults=defaults
        )
        
        if created:
            self.results['inserted'] += 1
        else:
            self.results['updated'] += 1

    def _convert_field_value(self, field_name: str, value):
        """
        Convertit la valeur en type approprié
        Gère aussi les ForeignKey en cherchant par 'nom'
        
        Args:
            field_name: Nom du champ
            value: Valeur brute du fichier Excel
            
        Returns:
            Valeur convertie appropriée
        """
        try:
            # Récupérer le champ du modèle
            field = self.Model._meta.get_field(field_name)
            field_type = field.get_internal_type()
            
            if field_type == 'ForeignKey':
                # Résoudre la relation par le champ 'nom' (cas courant)
                related_model = field.related_model
                try:
                    # D'abord essayer par 'nom'
                    return related_model.objects.get(nom=str(value).strip())
                except related_model.DoesNotExist:
                    # Si pas trouvé, essayer par ID
                    try:
                        return related_model.objects.get(id=int(value))
                    except (ValueError, related_model.DoesNotExist):
                        raise ValueError(f"Impossible de trouver {related_model.__name__} avec nom ou id='{value}'")
            
            elif field_type == 'ManyToManyField':
                # Gérer les M2M (pas supporté pour l'instant)
                raise ValueError("ManyToMany non supporté pour l'import")
            
            elif field_type in ['DateField', 'DateTimeField']:
                if isinstance(value, str):
                    try:
                        return datetime.strptime(value, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            return datetime.strptime(value, '%d/%m/%Y').date()
                        except ValueError:
                            raise ValueError(f"Format de date invalide: {value} (attendu: YYYY-MM-DD ou DD/MM/YYYY)")
                return value
            
            elif field_type == 'BooleanField':
                if isinstance(value, str):
                    return value.lower() in ['true', '1', 'yes', 'oui', 'yes']
                return bool(value)
            
            elif field_type in ['IntegerField', 'AutoField', 'BigIntegerField', 'SmallIntegerField']:
                return int(value)
            
            elif field_type in ['DecimalField', 'FloatField']:
                return float(value)
            
            return str(value).strip()
        except Exception as e:
            raise ValueError(f"Erreur conversion champ '{field_name}': {str(e)}")
